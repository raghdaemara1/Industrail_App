from openpyxl import Workbook
import os
import datetime
from config import OUTPUT_DIR

class SpreadsheetGenerator:
    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active) # Remove default sheet
        os.makedirs(OUTPUT_DIR, exist_ok=True)

    def create_workbook(self):
        pass

    def populate_rows(self, sheet_name: str, rows: list):
        ws = self.wb.create_sheet(title=sheet_name)
        if not rows:
            return
            
        columns = list(rows[0].keys())
        ws.append(columns)
        
        for row_dict in rows:
            row_data = []
            for col in columns:
                val = row_dict.get(col, "")
                row_data.append(val)
            ws.append(row_data)

        # Rule 1 from CLAUDE: Format Fault Code as text format "@"
        if "Fault Code *" in columns:
            col_idx = columns.index("Fault Code *") + 1
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = "@"

    def save(self, prefix: str) -> str:
        date_str = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        filename = f"{prefix}_{date_str}.xlsx"
        path = os.path.join(OUTPUT_DIR, filename)
        self.wb.save(path)
        return path
